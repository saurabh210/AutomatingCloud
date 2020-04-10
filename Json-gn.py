##This Script Converts excel data in JSON with 1st coloumn as ParameterKey and other coloums as Paramter Value
##A seprate JSON file will be created for mapping of 1st coloumn with each 'n' coloumns.
 ##Output JSON files will be stored from where you run the script.

#1. Please make sure there is no space left at start of any values in Excel data

#2. If your values in Excel contains integers , please make sure it starts with ' . 
#  For example, integer 100 should be written as '100 as it stores number as text in excel

##Please contact Saurabh Shah for any Enhancements/Modifications
try:
    import json
    from openpyxl import load_workbook
except ModuleNotFoundError:
     import os
     os.system('pip install openpyxl')
     import json
     from openpyxl import load_workbook

print("")
print("""
*******

This Script Converts excel data in JSON with 1st coloumn as ParameterKey and other coloums as Paramter Value
A seprate JSON file will be created for mapping of 1st coloumn with each 'n' coloumns.
 
Output JSON files will be stored from where you run the script.

1. Please make sure there is no space left at start of any values in Excel data

2. If your values in Excel contains integers , please make sure it starts with ' . 
   For example, integer 100 should be written as '100 as it stores number as text in excel.

*******
""")
print("")
#path = input("Please enter the path of Excel to execute: ")
wb=load_workbook('exceltojson.xlsx')
Sheet1 = wb.active
max_row=Sheet1.max_row
max_column=Sheet1.max_column
for i in range(2, max_column+1):
    listing = []
    
    for j in range(2, max_row+1):
         dictonery =  {}
         dictonery["ParameterKey"] = (Sheet1.cell(j,1)).value
         dictonery["ParameterValue"] = (Sheet1.cell(j,i)).value
         listing.append(dictonery)

    jsondump = json.dumps(listing,indent = 4, sort_keys=True)
    jsonfile = jsondump.replace('null','""')
    filename = (Sheet1.cell(5,i)).value + '.json'
    f = open(filename, 'w',encoding="UTF-8")
    f.write(jsonfile)
    f.close()