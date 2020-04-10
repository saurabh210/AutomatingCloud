##This Script adds tags (key and value) to All Volumes of EC2 listed in Excel.
##First Step is to set profile name of customer and add its aws access,secret,session key in credentials file and region in config file.
##Second Step is to edit "Data_TagVolumes.xlsx" and replace key , value and server names with required parameters
##Please contact Saurabh Shah for any Enhancements/Modifications

print("")
print("""
*******

This Script adds tags (key and value) to All Volumes of EC2 listed in Excel.
First Step is to set profile name of customer and add its aws access,secret,session key in credentials file and region in config file.
Second Step is to edit "Data_TagVolumes.xlsx" and replace key , value and server names with required parameters

*******
""")
print("")

try:
    import boto3,sys
    from openpyxl import load_workbook
except ModuleNotFoundError:
     import os
     os.system('pip install boto3 openpyxl')
     import boto3,sys
     from openpyxl import load_workbook

def taging(profilename):
    
    print("")
    print("Using Profile: {}".format(profilename))
    print("")
   # path = input("Please enter the full path of Excel to execute: ")
    wb=load_workbook('Data_TagVolumes.xlsx')
    Sheet1 = wb.active
    max_row=Sheet1.max_row
    max_column=Sheet1.max_column
    Tag_Key = Sheet1.cell(2,1).value
    Tag_Value = Sheet1.cell(2,2).value
    print("")
    print("All Volumes of each Below EC2 Servers will be Tagged:")
    print("")
    tag_vm = []
    for each in range(2,max_row+1):
        print("{}".format(Sheet1.cell(each,3).value))
        tag_vm.append(Sheet1.cell(each,3).value)
    print("")
    print("With Tag Key as : {}".format(Tag_Key))
    print("")
    print("And Tag Value as : {}".format(Tag_Value))
    print("")
    confirm = input("Please Confirm Yes to proceed or No to exit: ").lower()
    print("")
    if confirm == 'yes':
        session = boto3.Session(profile_name=profilename)        
        ec2 = session.client('ec2')
        tag_vol = []
        for vm in tag_vm: 
            response = ec2.describe_instances(
                Filters=[
                            {
                            'Name': 'tag:Name',
                            'Values': [vm]
                            },
                        ],
                    )
            for each in list(range(len(response['Reservations'][0]['Instances'][0]['BlockDeviceMappings']))):
                vm_vol = response['Reservations'][0]['Instances'][0]['BlockDeviceMappings'][each]['Ebs']['VolumeId']
                print("{}   {}".format(vm,vm_vol))
                tag_vol.append(vm_vol)
        ec2.create_tags(Resources= tag_vol,
            Tags=[
                    {
                    'Key': Tag_Key,
                    'Value': Tag_Value
                    },
                ]
            )
        print("")
        print("All Volumes Tagged. please validate on AWS Console")
        print("")
    else:
        print("")
        print ("OK Exited")
        print("")

def main():
    profilename = input("Please enter aws credentials profile name: ")
    taging(profilename)

if __name__ == '__main__':
    main()